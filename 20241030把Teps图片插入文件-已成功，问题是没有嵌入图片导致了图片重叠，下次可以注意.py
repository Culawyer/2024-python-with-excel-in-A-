import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# 定义文件夹路径和Excel文件路径
images_folder = 'D:\\您的新文件夹路径\折线图结果'  # 替换为你的图片文件夹路径
excel_file = 'D:\\投资\投资数据实验\跨表查询_沪深京股票(年频)roe 股东等\跨表查询_沪深京股票(年频).xlsx'  # 替换为你的Excel文件路径

# 加载Excel文件
workbook = load_workbook(excel_file)
sheet = workbook.active

# 获取所有图片文件
image_files = [f for f in os.listdir(images_folder) if f.endswith('.png')]

# 遍历所有图片文件
for image_file in image_files:
    # 获取前六个字符
    image_prefix = image_file[:6]

    # 在A列中查找对应的值
    rows_to_merge = []
    for row in range(2, sheet.max_row + 1):  # 假设数据从第二行开始
        cell_value = str(sheet[f'A{row}'].value)
        if cell_value == image_prefix:
            rows_to_merge.append(row)

    # 如果找到对应的行，插入图片
    if rows_to_merge:
        # 插入图片
        img = Image(os.path.join(images_folder, image_file))
        img.anchor = f'k{rows_to_merge[0]}'  # 将图片插入到k列第一个对应行
        sheet.add_image(img)

        # 合并单元格
        sheet.merge_cells(start_row=rows_to_merge[0], start_column=8,
                          end_row=rows_to_merge[-1], end_column=8)  # 合并H列

# 保存Excel文件
workbook.save(excel_file)
print("图片插入完成！")