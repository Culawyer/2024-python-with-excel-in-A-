import openpyxl

# A表和B表路径
a_table_path = r'D:\投资\投资数据实验\智能查询_沪深京股票(年频)2023eps\智能查询_沪深京股票(年频).xlsx'
b_table_path = r'D:\投资\投资数据实验\跨表查询_沪深京股票(年频)roe 股东等\跨表查询_沪深京股票(年频) - 副本.xlsx'

# 打开A表和B表
wb_a = openpyxl.load_workbook(a_table_path)
wb_b = openpyxl.load_workbook(b_table_path)

# 获取A表和B表的活动工作表
ws_a = wb_a.active
ws_b = wb_b.active

# 将A表中的证券代码和对应数据（F, G, H, I列）存储到字典中
data_to_insert = {}
for row in ws_a.iter_rows(min_row=2, min_col=1, max_col=9, values_only=True):
    sec_code = row[0]  # A列是证券代码
    data_to_insert[sec_code] = (row[5], row[6], row[7], row[8])  # F, G, H, I列数据

# 在B表中查找对应的证券代码，并将A表中的数据插入到B表的G, H, I, J列
for row in ws_b.iter_rows(min_row=2, min_col=1, max_col=14):  # 假设B表有14列，包含N列
    sec_code_b = row[0].value  # B表的A列是证券代码
    if sec_code_b in data_to_insert:
        f_value, g_value, h_value, i_value = data_to_insert[sec_code_b]

        # 检查G, H, I, J列是否为空，且检查N列是否已被覆盖
        if ws_b.cell(row=row[6].row, column=7).value is None and ws_b.cell(row=row[6].row, column=15).value is None:  # G列和N列
            ws_b.cell(row=row[6].row, column=7, value=f_value)
        if ws_b.cell(row=row[7].row, column=8).value is None and ws_b.cell(row=row[7].row, column=15).value is None:  # H列和N列
            ws_b.cell(row=row[7].row, column=8, value=g_value)
        if ws_b.cell(row=row[8].row, column=9).value is None and ws_b.cell(row=row[8].row, column=15).value is None:  # I列和N列
            ws_b.cell(row=row[8].row, column=9, value=h_value)
        if ws_b.cell(row=row[9].row, column=10).value is None and ws_b.cell(row=row[9].row, column=15).value is None:  # J列和N列
            ws_b.cell(row=row[9].row, column=10, value=i_value)

# 保存修改后的B表
output_path = r'D:\投资\投资数据实验\跨表查询_沪深京股票(年频)roe 股东等\跨表查询_沪深京股票(年频)更新.xlsx'
wb_b.save(output_path)

print("数据插入完成，避免覆盖已有图片或内容的单元格。")

# 评价：虽然还是覆盖了但是格式和对照上没有问题，所以直接插入复制了，这样也比较轻松。
# 需要注意每一版文件的备份，非常必要。（不光是保存，而是备份。）
